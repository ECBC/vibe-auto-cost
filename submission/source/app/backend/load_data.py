"""
Data loading, Heuristic Alignment Matrix, and cost calculation engine.
Loads the European Cars xlsx + UCI car.data at import time into in-memory indices.
"""

import os
import csv
import statistics
from pathlib import Path
from typing import Optional

import openpyxl

# ---------------------------------------------------------------------------
# Paths
# ---------------------------------------------------------------------------
BASE_DIR = Path(__file__).resolve().parent.parent.parent  # /workspace/auto-evaluation
DATA_DIR = BASE_DIR / "data"
EUROPEAN_XLSX = DATA_DIR / "extracted_european" / "EUROPEAN CARS DATASET.xlsx"
UCI_CAR_DATA = DATA_DIR / "extracted_uci" / "car.data"

# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------
MAINT_COST_MAP = {"vhigh": 1200, "high": 850, "med": 500, "low": 250}
PREMIUM_BRANDS = {
    "bmw", "mercedes-benz", "mercedes", "audi", "porsche", "lexus",
    "jaguar", "maserati", "bentley", "ferrari", "lamborghini",
    "aston martin", "rolls-royce", "tesla", "ds automobiles",
}
ELECTRICITY_RATE = 0.30  # €/kWh
ANNUAL_KM = 15000

# ---------------------------------------------------------------------------
# In-memory stores (populated by load_all())
# ---------------------------------------------------------------------------
european_records: list[dict] = []
european_by_brand: dict[str, list[dict]] = {}
european_by_brand_model: dict[tuple[str, str], list[dict]] = {}
uci_records: list[dict] = []
global_median_fuel: float = 0.0
global_median_maint: float = 0.0
all_brands: list[str] = []
models_by_brand: dict[str, list[str]] = {}
engines_by_brand_model: dict[str, list[str]] = {}


def _extract_if_missing():
    """Re-extract zips if the extracted directories are missing (reproducibility)."""
    euro_zip = DATA_DIR / "archive.zip"
    uci_zip = DATA_DIR / "car+evaluation.zip"
    if not EUROPEAN_XLSX.exists() and euro_zip.exists():
        import zipfile
        with zipfile.ZipFile(euro_zip, "r") as z:
            z.extractall(DATA_DIR / "extracted_european")
    if not UCI_CAR_DATA.exists() and uci_zip.exists():
        import zipfile
        with zipfile.ZipFile(uci_zip, "r") as z:
            z.extractall(DATA_DIR / "extracted_uci")


def _load_european():
    """Load the European Cars xlsx into memory."""
    global european_records, european_by_brand, european_by_brand_model
    global global_median_fuel, global_median_maint, all_brands, models_by_brand, engines_by_brand_model

    wb = openpyxl.load_workbook(str(EUROPEAN_XLSX), read_only=True, data_only=True)
    ws = wb.active
    headers = [cell.value for cell in ws[1]]

    for row in ws.iter_rows(min_row=2, values_only=True):
        rec = dict(zip(headers, row))
        if not rec.get("Brand"):
            continue
        european_records.append(rec)
        brand_lower = rec["Brand"].strip().lower()
        european_by_brand.setdefault(brand_lower, []).append(rec)
        model_name = (rec.get("Model Name") or "").strip().lower()
        european_by_brand_model.setdefault((brand_lower, model_name), []).append(rec)

    wb.close()

    # Compute global medians for fallback
    fuel_costs = [r["Energy Cost (€/100km)"] for r in european_records if r.get("Energy Cost (€/100km)") and r["Energy Cost (€/100km)"] > 0]
    maint_costs = [r["Maintenance Cost (€/year)"] for r in european_records if r.get("Maintenance Cost (€/year)") and r["Maintenance Cost (€/year)"] > 0]
    global_median_fuel = statistics.median(fuel_costs) if fuel_costs else 8.0
    global_median_maint = statistics.median(maint_costs) if maint_costs else 500.0

    # Build dropdown data
    brand_set = sorted(set(r["Brand"].strip() for r in european_records if r.get("Brand")))
    all_brands.extend(brand_set)
    for rec in european_records:
        brand = rec["Brand"].strip()
        model = rec.get("Model Name", "").strip()
        fuel = rec.get("Fuel Type", "").strip()
        if model:
            models_by_brand.setdefault(brand, [])
            if model not in models_by_brand[brand]:
                models_by_brand[brand].append(model)
        if model and fuel:
            key = f"{brand}|{model}"
            engines_by_brand_model.setdefault(key, [])
            if fuel not in engines_by_brand_model[key]:
                engines_by_brand_model[key].append(fuel)


def _load_uci():
    """Load UCI car.data (unheadered CSV)."""
    global uci_records
    UCI_COLS = ["buying", "maint", "doors", "persons", "lug_boot", "safety", "class"]
    with open(UCI_CAR_DATA, "r") as f:
        reader = csv.reader(f)
        for row in reader:
            if len(row) == 7:
                uci_records.append(dict(zip(UCI_COLS, [v.strip() for v in row])))


def load_all():
    """Load all datasets. Called once at app startup."""
    _extract_if_missing()
    _load_european()
    _load_uci()


# ---------------------------------------------------------------------------
# Heuristic Alignment Matrix
# ---------------------------------------------------------------------------

def _get_buying_tier(make: str) -> str:
    """Map make to UCI buying tier."""
    if make.lower() in PREMIUM_BRANDS:
        return "vhigh"
    return "med"


def _get_maint_tier(engine: str, horsepower: Optional[float] = None) -> str:
    """Map engine type to UCI maint tier."""
    engine_lower = engine.lower() if engine else ""
    if "electric" in engine_lower:
        return "high"
    if "hybrid" in engine_lower or "electric" in engine_lower:
        return "med"
    # Use horsepower as proxy for engine size
    if horsepower and horsepower > 200:
        return "high"
    if horsepower and horsepower < 120:
        return "low"
    return "med"


def _find_uci_matches(buying_tier: str, maint_tier: str) -> list[dict]:
    """Find UCI records matching the buying and maint tiers."""
    matches = [r for r in uci_records if r["buying"] == buying_tier and r["maint"] == maint_tier]
    if not matches:
        # Relax: match just buying tier
        matches = [r for r in uci_records if r["buying"] == buying_tier]
    if not matches:
        matches = uci_records  # all as fallback
    return matches


def _strip_model(model: str) -> str:
    """Strip sub-model badges to get base model name."""
    # Remove common suffixes like GTI, TDI, etc.
    parts = model.strip().split()
    if len(parts) > 1:
        # Return first 1-2 significant words
        return parts[0]
    return model.strip()


def _find_european_match(make: str, model: str, engine: str) -> tuple[list[dict], str]:
    """
    Find European dataset records for given make/model/engine.
    Returns (matching_records, match_type).
    Fallback chain: exact model → stripped model → brand + engine → brand → global median.
    """
    brand_lower = make.strip().lower()
    model_lower = model.strip().lower()
    engine_lower = engine.strip().lower() if engine else ""

    # 1. Exact brand + model match
    key = (brand_lower, model_lower)
    if key in european_by_brand_model:
        recs = european_by_brand_model[key]
        # Filter by engine/fuel if possible
        fuel_filtered = [r for r in recs if r.get("Fuel Type", "").lower() == engine_lower]
        if fuel_filtered:
            return fuel_filtered, "exact"
        return recs, "exact_no_fuel"

    # 2. Strip model and try again
    stripped = _strip_model(model_lower)
    for k, v in european_by_brand_model.items():
        if k[0] == brand_lower and stripped in k[1]:
            fuel_filtered = [r for r in v if r.get("Fuel Type", "").lower() == engine_lower]
            if fuel_filtered:
                return fuel_filtered, "stripped"
            return v, "stripped_no_fuel"

    # 3. Brand match with engine filter
    if brand_lower in european_by_brand:
        brand_recs = european_by_brand[brand_lower]
        fuel_filtered = [r for r in brand_recs if r.get("Fuel Type", "").lower() == engine_lower]
        if fuel_filtered:
            return fuel_filtered, "brand_fuel"
        return brand_recs, "brand_only"

    # 4. Global fallback (empty list triggers median usage)
    return [], "global_fallback"


# ---------------------------------------------------------------------------
# Cost Calculation Engine
# ---------------------------------------------------------------------------

def calculate_evaluation(make: str, model: str, engine: str) -> dict:
    """
    Main evaluation function. Returns the full response dict
    conforming to the logic_guide §4 schema.
    """
    euro_matches, match_type = _find_european_match(make, model, engine)

    # Determine buying/maint tiers for UCI alignment
    buying_tier = _get_buying_tier(make)

    # Get average horsepower from matches for maint tier heuristic
    avg_hp = None
    if euro_matches:
        hps = [r["Horsepower (HP)"] for r in euro_matches if r.get("Horsepower (HP)") and r["Horsepower (HP)"] > 0]
        if hps:
            avg_hp = statistics.mean(hps)

    maint_tier = _get_maint_tier(engine, avg_hp)

    # UCI alignment
    uci_matches = _find_uci_matches(buying_tier, maint_tier)

    # --- Financial Projections ---
    # Annual fuel cost
    if euro_matches:
        fuel_values = [r["Energy Cost (€/100km)"] for r in euro_matches if r.get("Energy Cost (€/100km)") and r["Energy Cost (€/100km)"] > 0]
        if fuel_values:
            energy_cost_per_100km = statistics.mean(fuel_values)
        else:
            energy_cost_per_100km = global_median_fuel
    else:
        energy_cost_per_100km = global_median_fuel

    annual_fuel_cost = energy_cost_per_100km * (ANNUAL_KM / 100)

    # For EV: use Efficiency (Wh/km) if available
    if engine and "electric" in engine.lower() and euro_matches:
        eff_values = [r["Efficiency (Wh/km)"] for r in euro_matches if r.get("Efficiency (Wh/km)") and r["Efficiency (Wh/km)"] > 0]
        if eff_values:
            avg_eff = statistics.mean(eff_values)
            annual_fuel_cost = avg_eff * ANNUAL_KM / 1000 * ELECTRICITY_RATE

    # Annual maintenance cost (UCI-mapped primary, dataset fallback)
    annual_maintenance_cost = float(MAINT_COST_MAP.get(maint_tier, 500))

    # Total 5-year
    total_5_year_cost = (annual_fuel_cost + annual_maintenance_cost) * 5

    # --- Guardrails ---
    flagged = False
    if total_5_year_cost > 150000:
        total_5_year_cost = 150000.0
        flagged = True

    # --- Reliability Scores ---
    # Safety from European dataset
    if euro_matches:
        safety_vals = [r["Safety Rating (Euro NCAP)"] for r in euro_matches if r.get("Safety Rating (Euro NCAP)")]
        avg_safety = statistics.mean([float(s) for s in safety_vals if s]) if safety_vals else 3.0
    else:
        avg_safety = 3.0

    safety_labels = {5: "Excellent", 4: "Good", 3: "Average", 2: "Below Average", 1: "Poor", 0: "Not Rated"}
    safety_rating = safety_labels.get(round(avg_safety), "Average")

    # Comfort rating from boot capacity + seating + ADAS
    if euro_matches:
        boot_vals = [r["Boot Capacity (L)"] for r in euro_matches if r.get("Boot Capacity (L)") and r["Boot Capacity (L)"] > 0]
        seat_vals = [r["Seating Capacity"] for r in euro_matches if r.get("Seating Capacity") and r["Seating Capacity"] > 0]
        avg_boot = statistics.mean(boot_vals) if boot_vals else 350
        avg_seats = statistics.mean(seat_vals) if seat_vals else 5
    else:
        avg_boot = 350
        avg_seats = 5

    comfort_score = min(100, (avg_boot / 600 * 40) + (avg_seats / 7 * 30) + (avg_safety / 5 * 30))
    if comfort_score >= 75:
        comfort_rating = "Excellent"
    elif comfort_score >= 55:
        comfort_rating = "Good"
    elif comfort_score >= 35:
        comfort_rating = "Average"
    else:
        comfort_rating = "Below Average"

    # Overall vibe score: combines safety + UCI class distribution + comfort
    uci_class_scores = {"vgood": 100, "good": 75, "acc": 50, "unacc": 25}
    if uci_matches:
        class_vals = [uci_class_scores.get(r["class"], 50) for r in uci_matches]
        avg_class = statistics.mean(class_vals)
    else:
        avg_class = 50

    overall_vibe_score = int(round(
        (avg_safety / 5 * 35) +  # 35% safety
        (avg_class / 100 * 35) +  # 35% UCI class
        (comfort_score / 100 * 30)  # 30% comfort
    ))
    # Clamp [0, 100]
    overall_vibe_score = max(0, min(100, overall_vibe_score))

    # --- Radar chart data (for frontend) ---
    # Map UCI safety distribution
    safety_map = {"high": 5, "med": 3, "low": 1}
    uci_safety_vals = [safety_map.get(r["safety"], 3) for r in uci_matches] if uci_matches else [3]
    radar_safety = statistics.mean(uci_safety_vals) / 5 * 100

    # Maintenance (inverted: low maint = high score)
    maint_inv_map = {"low": 100, "med": 65, "high": 35, "vhigh": 10}
    radar_maint = maint_inv_map.get(maint_tier, 50)

    # Boot capacity
    lug_map = {"big": 90, "med": 60, "small": 30}
    uci_lug_vals = [lug_map.get(r["lug_boot"], 50) for r in uci_matches] if uci_matches else [50]
    radar_boot = statistics.mean(uci_lug_vals)

    # Comfort (reuse comfort_score)
    radar_comfort = comfort_score

    return {
        "status": "success",
        "metadata": {
            "requested_make": make,
            "requested_model": model,
            "matched_records": len(euro_matches),
            "flagged": flagged,
        },
        "reliability_scores": {
            "safety_rating": safety_rating,
            "comfort_rating": comfort_rating,
            "overall_vibe_score": overall_vibe_score,
        },
        "financial_projections": {
            "annual_fuel_cost": round(annual_fuel_cost, 2),
            "annual_maintenance_cost": round(annual_maintenance_cost, 2),
            "total_5_year_cost": round(total_5_year_cost, 2),
        },
        "radar_data": {
            "safety": round(radar_safety, 1),
            "maintenance": round(radar_maint, 1),
            "boot_capacity": round(radar_boot, 1),
            "comfort": round(radar_comfort, 1),
        },
        "yearly_breakdown": [
            {
                "year": y,
                "cumulative_fuel": round(annual_fuel_cost * y, 2),
                "cumulative_maintenance": round(annual_maintenance_cost * y, 2),
            }
            for y in range(1, 6)
        ],
    }
